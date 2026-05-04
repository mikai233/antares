import os
from pathlib import Path
from shutil import copyfile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATAS = ROOT / "Datas"
DATAS.mkdir(parents=True, exist_ok=True)
EXAMPLES_ROOT = Path(os.environ.get("LUBAN_EXAMPLES_ROOT", "/tmp/luban_examples"))
MINI_DATAS = EXAMPLES_ROOT / "MiniTemplate" / "Datas"
FULL_DATAS = EXAMPLES_ROOT / "DataTables" / "Datas"


def require(path: Path) -> Path:
    if not path.exists():
        raise SystemExit(f"required template file not found: {path}")
    return path


def clear_sheet(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    for r in range(4, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None
    return wb, ws


copyfile(require(MINI_DATAS / "__tables__.xlsx"), DATAS / "__tables__.xlsx")
wb, ws = clear_sheet(DATAS / "__tables__.xlsx")
table_rows = [
    ("", "game.TbRotationMessage", "RotationMessage", True, "#game.rotation_message.xlsx", None, "list", "s", "轮播消息列表", None, None),
    ("", "game.TbGameGlobal", "GameGlobal", True, "#game.game_global.xlsx", None, "one", "s", "全局单例配置", None, None),
]
for i, row in enumerate(table_rows, start=4):
    for j, value in enumerate(row, start=1):
        ws.cell(i, j).value = value
wb.save(DATAS / "__tables__.xlsx")

copyfile(require(MINI_DATAS / "__beans__.xlsx"), DATAS / "__beans__.xlsx")
wb, ws = clear_sheet(DATAS / "__beans__.xlsx")
bean_rows = [
    ("", "game.Reward", None, None, None, None, None, None, None, "itemId", "物品ID", "int", None, "物品ID", None, None),
    ("", None, None, None, None, None, None, None, None, "count", "数量", "int", None, "数量", None, None),
    ("", "game.DropEntry", None, None, None, None, None, None, None, "itemId", "物品ID", "int", None, "物品ID", None, None),
    ("", None, None, None, None, None, None, None, None, "minCount", "最小数量", "int", None, "最小数量", None, None),
    ("", None, None, None, None, None, None, None, None, "maxCount", "最大数量", "int", None, "最大数量", None, None),
    ("", None, None, None, None, None, None, None, None, "weight", "权重", "int", None, "权重", None, None),
    ("", "game.Vec2", None, None, None, None, None, None, None, "x", "X", "int", None, "X", None, None),
    ("", None, None, None, None, None, None, None, None, "y", "Y", "int", None, "Y", None, None),
    ("", "game.Rect", None, None, None, None, None, None, None, "x", "X", "int", None, "X", None, None),
    ("", None, None, None, None, None, None, None, None, "y", "Y", "int", None, "Y", None, None),
    ("", None, None, None, None, None, None, None, None, "width", "宽", "int", None, "宽", None, None),
    ("", None, None, None, None, None, None, None, None, "height", "高", "int", None, "高", None, None),
]
for i, row in enumerate(bean_rows, start=4):
    for j, value in enumerate(row, start=1):
        ws.cell(i, j).value = value
wb.save(DATAS / "__beans__.xlsx")

copyfile(require(FULL_DATAS / "__enums__.xlsx"), DATAS / "__enums__.xlsx")
wb, ws = clear_sheet(DATAS / "__enums__.xlsx")
rows = [
    ("", "item.ItemType", False, True, None, None, None, "Currency", "货币", 1, "货币", None),
    ("", None, None, None, None, None, None, "Consumable", "消耗品", 2, "消耗品", None),
    ("", None, None, None, None, None, None, "Equipment", "装备", 3, "装备", None),
    ("", None, None, None, None, None, None, "Material", "材料", 4, "材料", None),
]
for i, row in enumerate(rows, start=4):
    for j, value in enumerate(row, start=1):
        ws.cell(i, j).value = value
wb.save(DATAS / "__enums__.xlsx")

copyfile(require(MINI_DATAS / "#demo.item.xlsx"), DATAS / "#game.item.xlsx")
wb = load_workbook(DATAS / "#game.item.xlsx")
ws = wb.active
for r in range(1, max(ws.max_row, 8) + 5):
    for c in range(1, 12):
        ws.cell(r, c).value = None
rows = [
    ("##var", "id", "name", "type", "quality", "maxStack", "sellPrice"),
    ("##type", "int", "string", "item.ItemType", "int", "int", "int"),
    ("##group", "s", "s", "s", "s", "s", "s"),
    ("##", "编号", "名称", "类型", "品质", "最大堆叠", "出售价格"),
    ("", 1001, "Gold", "Currency", 1, 999999, 0),
    ("", 2001, "Small Potion", "Consumable", 2, 99, 10),
    ("", 3001, "Iron Sword", "Equipment", 3, 1, 120),
    ("", 4001, "Wolf Fang", "Material", 2, 999, 3),
]
for i, row in enumerate(rows, start=1):
    for j, value in enumerate(row, start=1):
        ws.cell(i, j).value = value
wb.save(DATAS / "#game.item.xlsx")

copyfile(require(MINI_DATAS / "#demo.item.xlsx"), DATAS / "#game.activity.xlsx")
wb = load_workbook(DATAS / "#game.activity.xlsx")
ws = wb.active
for r in range(1, max(ws.max_row, 6) + 5):
    for c in range(1, 12):
        ws.cell(r, c).value = None
rows = [
    ("##var", "id", "name", "startTime", "endTime", "unlockLevel", "conditionSummary", "rewardSummary"),
    ("##type", "string", "string", "string", "string", "int", "string", "string"),
    ("##group", "s", "s", "s", "s", "s", "s", "s"),
    ("##", "编号", "名称", "开始时间", "结束时间", "解锁等级", "条件摘要", "奖励摘要"),
    ("", "daily_login", "Daily Login", "2026-01-01T00:00:00", "2026-12-31T23:59:59", 1, "loginDays=1", "1001x100,2001x1"),
    ("", "wolf_hunt", "Wolf Hunt", "2026-05-01T00:00:00", "2026-05-14T23:59:59", 5, "killCount=10,monsterId=101", "1001x500,3001x1"),
]
for i, row in enumerate(rows, start=1):
    for j, value in enumerate(row, start=1):
        ws.cell(i, j).value = value
wb.save(DATAS / "#game.activity.xlsx")

copyfile(require(MINI_DATAS / "#demo.item.xlsx"), DATAS / "#game.monster.xlsx")
wb = load_workbook(DATAS / "#game.monster.xlsx")
ws = wb.active
for r in range(1, max(ws.max_row, 7) + 5):
    for c in range(1, 16):
        ws.cell(r, c).value = None
rows = [
    ("##var", "id", "name", "level", "hp", "sceneId", "skillIds", "rewards"),
    ("##type", "int", "string", "int", "long", "int", "(array#sep=,),int", "(list#sep=,),game.Reward"),
    ("##group", "s", "s", "s", "s", "s", "s", "s"),
    ("##", "编号", "名称", "等级", "生命值", "场景ID", "技能ID列表", "奖励列表"),
    ("", 101, "Forest Wolf", 5, 1200, 1, "1001,1002", "4001,2,1001,50"),
    ("", 201, "Cave Troll", 12, 9800, 2, "2001,2002,2003", "4001,8,3001,1"),
]
for i, row in enumerate(rows, start=1):
    for j, value in enumerate(row, start=1):
        ws.cell(i, j).value = value
wb.save(DATAS / "#game.monster.xlsx")

drop_pool_legacy = DATAS / "#game.drop_pool.xlsx"
if drop_pool_legacy.exists():
    drop_pool_legacy.unlink()

copyfile(require(MINI_DATAS / "#demo.item.xlsx"), DATAS / "#game.droppool.xlsx")
wb = load_workbook(DATAS / "#game.droppool.xlsx")
ws = wb.active
for r in range(1, max(ws.max_row, 7) + 5):
    for c in range(1, 16):
        ws.cell(r, c).value = None
rows = [
    ("##var", "id", "rolls", "entries"),
    ("##type", "int", "int", "(list#sep=,),game.DropEntry"),
    ("##group", "s", "s", "s"),
    ("##", "编号", "抽取次数", "掉落项"),
    ("", 1, 2, "1001,100,200,80,2001,1,2,15,3001,1,1,5"),
    ("", 2, 1, "4001,2,5,70,3001,1,1,30"),
]
for i, row in enumerate(rows, start=1):
    for j, value in enumerate(row, start=1):
        ws.cell(i, j).value = value
wb.save(DATAS / "#game.droppool.xlsx")

copyfile(require(MINI_DATAS / "#demo.item.xlsx"), DATAS / "#game.scene.xlsx")
wb = load_workbook(DATAS / "#game.scene.xlsx")
ws = wb.active
for r in range(1, max(ws.max_row, 7) + 5):
    for c in range(1, 16):
        ws.cell(r, c).value = None
rows = [
    ("##var", "id", "name", "width", "height", "spawnPoints", "safeZones"),
    ("##type", "int", "string", "int", "int", "(list#sep=,),game.Vec2", "(list#sep=,),game.Rect"),
    ("##group", "s", "s", "s", "s", "s", "s"),
    ("##", "编号", "名称", "宽", "高", "出生点", "安全区"),
    ("", 1, "Novice Plains", 100, 80, "10,10,20,20,30,30", "0,0,12,12,80,60,20,20"),
    ("", 2, "Troll Cave", 60, 40, "5,5,15,10", "0,0,8,8"),
]
for i, row in enumerate(rows, start=1):
    for j, value in enumerate(row, start=1):
        ws.cell(i, j).value = value
wb.save(DATAS / "#game.scene.xlsx")

copyfile(require(MINI_DATAS / "#demo.item.xlsx"), DATAS / "#game.rotation_message.xlsx")
wb = load_workbook(DATAS / "#game.rotation_message.xlsx")
ws = wb.active
for r in range(1, max(ws.max_row, 7) + 5):
    for c in range(1, 8):
        ws.cell(r, c).value = None
rows = [
    ("##var", "id", "content", "minLevel"),
    ("##type", "int", "string", "int"),
    ("##group", "s", "s", "s"),
    ("##", "编号", "内容", "最低等级"),
    ("", 1, "Welcome to Antares", 1),
    ("", 2, "Weekend bonus is live", 10),
    ("", 3, "World boss opens at 20:00", 20),
]
for i, row in enumerate(rows, start=1):
    for j, value in enumerate(row, start=1):
        ws.cell(i, j).value = value
wb.save(DATAS / "#game.rotation_message.xlsx")

copyfile(require(MINI_DATAS / "#demo.item.xlsx"), DATAS / "#game.game_global.xlsx")
wb = load_workbook(DATAS / "#game.game_global.xlsx")
ws = wb.active
for r in range(1, max(ws.max_row, 6) + 5):
    for c in range(1, 8):
        ws.cell(r, c).value = None
rows = [
    ("##var", "defaultWorldId", "maxPlayerLevel", "maintenanceNotice"),
    ("##type", "int", "int", "string"),
    ("##group", "s", "s", "s"),
    ("##", "默认世界ID", "最大玩家等级", "维护公告"),
    ("", 1, 60, "No maintenance scheduled"),
]
for i, row in enumerate(rows, start=1):
    for j, value in enumerate(row, start=1):
        ws.cell(i, j).value = value
wb.save(DATAS / "#game.game_global.xlsx")

print(f"Generated workbooks under {DATAS}")
